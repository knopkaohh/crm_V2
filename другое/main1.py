import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import math

# Сегментированная модель v4
def calculate_price(width, length, quantity):
    if quantity <= 0:
        return 0
    log_q = math.log(quantity)
    if width <= 20:  # Small
        base = 0.6806 + 0.1103 * width + 0.0199 * length - 0.1710 * log_q
    elif width <= 40:  # Mid
        base = 4.5297 + 0.0405 * width + 0.0058 * length - 0.3704 * log_q
    else:  # Big
        base = 8.0779 - 0.0090 * width + 0.0156 * length - 0.6160 * log_q
    return round(base, 2)

# Создаём workbook
wb = openpyxl.Workbook()

# Лист 1: Прайс-лист
ws1 = wb.active
ws1.title = "Прайс-лист"

# Заголовки
ws1['A1'] = "Размер (мм)"
quantities = [1000, 2000, 3000, 4000, 5000, 10000]
for col, q in enumerate(quantities, start=2):
    ws1.cell(row=1, column=col, value=f"{q} шт.")
    ws1.cell(row=1, column=col).font = Font(bold=True)
    ws1.cell(row=1, column=col).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Генерируем размеры
widths = list(range(10, 85, 5))
lengths = list(range(30, 155, 5))
row = 2
for w in widths:
    for l in lengths:
        size = f"{w}x{l}"
        ws1.cell(row=row, column=1, value=size)
        for col, q in enumerate(quantities, start=2):
            price = calculate_price(w, l, q)
            ws1.cell(row=row, column=col, value=price)
        row += 1

# Автоподгонка колонок
for col in ws1.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 15)
    ws1.column_dimensions[column].width = adjusted_width

# Лист 2: Калькулятор
ws2 = wb.create_sheet("Калькулятор")

# Форма
ws2['A3'] = "Ширина (мм):"
ws2['B3'] = 15  # Пример
ws2['A4'] = "Длина (мм):"
ws2['B4'] = 60  # Пример
ws2['A5'] = "Количество (шт.):"
ws2['B5'] = 1000  # Пример

ws2['A7'] = "Цена за единицу (руб.):"
# Сегментированная формула в Excel (IF для групп)
ws2['B7'] = """=IF(B3<=20, ROUND(0.6806 + 0.1103*B3 + 0.0199*B4 - 0.1710*LN(B5), 2),
IF(B3<=40, ROUND(4.5297 + 0.0405*B3 + 0.0058*B4 - 0.3704*LN(B5), 2),
ROUND(8.0779 - 0.0090*B3 + 0.0156*B4 - 0.6160*LN(B5), 2)))"""

ws2['A9'] = "Общая сумма (руб.):"
ws2['B9'] = "=B7 * B5"

# Стили
header_font = Font(bold=True)
for cell in ['A3', 'A4', 'A5', 'A7', 'A9']:
    ws2[cell].font = header_font
    ws2[cell].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

# Инструкция
ws2['A11'] = "Инструкция: Впиши в B3-B5. Модель v4 сегментирована по ширине (R²~0.75 общая)."
ws2['A11'].font = Font(italic=True)

# Сохраняем
wb.save("прайс_бирки_v4.xlsx")
print("Файл 'прайс_бирки_v4.xlsx' создан успешно!")